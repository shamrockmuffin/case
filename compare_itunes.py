import csv
from openpyxl import load_workbook
from datetime import datetime

def load_csv_log(filename):
    calls = {}
    with open(filename, 'r') as file:
        reader = csv.reader(file)
        next(reader)  # Skip header
        for row in reader:
            if not row:  # Skip empty rows
                continue
            call_id = row[0]
            calls[call_id] = {
                'id': call_id,
                'phone': row[1],
                'count': row[2],
                'timestamp': row[4],
                'type': row[6],
                'battery': row[8],
                'area_code': row[10]
            }
    return calls

def load_excel_log(filename):
    calls = {}
    wb = load_workbook(filename)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    
    for row in ws.iter_rows(min_row=2):
        values = [cell.value for cell in row]
        row_data = dict(zip(headers, values))
        
        # Create a unique ID for Excel records
        call_id = f"{row_data.get('Phone Number', '')}_{row_data.get('Timestamp', '')}"
        calls[call_id] = {
            'id': call_id,
            'phone': str(row_data.get('Phone Number', '')),
            'timestamp': str(row_data.get('Timestamp', '')),
            'type': str(row_data.get('Call Type', '')),
            'area_code': str(row_data.get('Area Code', ''))
        }
    return calls

# Load both logs
print("Loading CSV log...")
csv_log = load_csv_log('call_history.csv')
print("Loading iTunes log...")
excel_log = load_excel_log('Itunes-call-history.xlsx')

# Create sets of phone numbers and timestamps for comparison
csv_calls = {(call['phone'], call['timestamp']) for call in csv_log.values()}
excel_calls = {(call['phone'], call['timestamp']) for call in excel_log.values()}

# Find unique calls in each log
csv_only = csv_calls - excel_calls
excel_only = excel_calls - csv_calls
common = csv_calls & excel_calls

# Print results
print("\nCall Log Comparison Analysis")
print("===========================")
print(f"\nTotal calls in CSV: {len(csv_log)}")
print(f"Total calls in iTunes: {len(excel_log)}")
print(f"Common calls: {len(common)}")

print("\nCalls only in CSV file (first 10):")
for phone, timestamp in sorted(list(csv_only)[:10], key=lambda x: x[1]):
    print(f"- {timestamp}: {phone}")

print("\nCalls only in iTunes file (first 10):")
for phone, timestamp in sorted(list(excel_only)[:10], key=lambda x: x[1]):
    print(f"- {timestamp}: {phone}")

# Analyze call patterns
csv_numbers = {}
excel_numbers = {}

for call in csv_log.values():
    phone = call['phone']
    if phone not in csv_numbers:
        csv_numbers[phone] = 0
    csv_numbers[phone] += 1

for call in excel_log.values():
    phone = call['phone']
    if phone not in excel_numbers:
        excel_numbers[phone] = 0
    excel_numbers[phone] += 1

print("\nTop 5 most called numbers in CSV:")
for phone, count in sorted(csv_numbers.items(), key=lambda x: x[1], reverse=True)[:5]:
    print(f"- {phone}: {count} calls")

print("\nTop 5 most called numbers in iTunes:")
for phone, count in sorted(excel_numbers.items(), key=lambda x: x[1], reverse=True)[:5]:
    print(f"- {phone}: {count} calls") 
